"""导入 UCI Online Retail II 真实交易数据 → 映射到业务表。

用法 (容器内): python -m app.seed.import_uci
- 下载 online_retail_II.xlsx (UCI), 流式解析 (read_only)
- 日期重定基: 把数据最新日期对齐到"今天", 使「近 7 天」等时间查询有数据
- 映射: Invoice→Order, 行→OrderItem, StockCode→Product, Customer ID→Customer,
  负数量/以 C 开头的发票→Return; 库存/广告/工单按规则合成
- 保留演示用异常 SKU NX-AIR-FRYER-001 与订单 10086
- 为控制本地负载, 默认上限 MAX_ORDERS 张发票 (不静默截断, 会打印)
"""
from __future__ import annotations

import os
import random
from datetime import datetime, timedelta, timezone

import httpx
from openpyxl import load_workbook
from sqlalchemy import text

from app.db.engine import SessionLocal
from app.db.models import (
    Campaign,
    Customer,
    Inventory,
    Order,
    OrderItem,
    Product,
    Return,
    SupportTicket,
)

URL = "https://archive.ics.uci.edu/ml/machine-learning-databases/00502/online_retail_II.xlsx"
XLSX = "/tmp/online_retail_II.xlsx"
SHEET = "Year 2010-2011"
DATA_MAX = datetime(2011, 12, 9, tzinfo=timezone.utc)  # 数据集最新日期 (近似)
NOW = datetime.now(timezone.utc)
OFFSET = NOW - DATA_MAX
MAX_ORDERS = 4000

ANOMALY_SKU = "NX-AIR-FRYER-001"
random.seed(42)


def download() -> None:
    if os.path.exists(XLSX) and os.path.getsize(XLSX) > 1_000_000:
        print(f"已存在 {XLSX}")
        return
    print(f"下载 {URL} ...")
    with httpx.stream("GET", URL, timeout=120, follow_redirects=True) as r:
        r.raise_for_status()
        with open(XLSX, "wb") as f:
            for chunk in r.iter_bytes(1 << 20):
                f.write(chunk)
    print(f"下载完成 {os.path.getsize(XLSX)//1024} KB")


def _rebase(dt) -> datetime:
    if dt is None:
        return NOW
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt + OFFSET


def truncate(db) -> None:
    db.execute(text(
        "TRUNCATE products, customers, orders, order_items, inventory, returns, "
        "campaigns, support_tickets RESTART IDENTITY CASCADE"
    ))
    db.commit()


def run_import(db) -> None:
    download()
    print("解析 xlsx (read_only)...")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[-1]]

    products: dict[str, Product] = {}
    customers: dict[str, int] = {}
    orders: dict[str, Order] = {}
    items: list[OrderItem] = []
    returns: list[Return] = []
    next_order_id = 1
    next_cust_id = 1
    capped = False

    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        # Invoice, StockCode, Description, Quantity, InvoiceDate, Price, Customer ID, Country
        invoice, stock, desc, qty, idate, price, cust, _country = row[:8]
        if stock is None or price is None:
            continue
        sku = str(stock).strip()
        try:
            qty = int(qty or 0)
            price = float(price or 0)
        except (TypeError, ValueError):
            continue
        created = _rebase(idate)

        # product (lazy)
        if sku not in products:
            products[sku] = Product(sku=sku, name=(str(desc).strip()[:255] if desc else sku),
                                    category="retail", price=round(abs(price), 2), status="active")

        is_return = (invoice and str(invoice).upper().startswith("C")) or qty < 0
        if is_return:
            returns.append(Return(order_id=0, sku=sku, reason="customer return",
                                  amount=round(abs(qty * price), 2), created_at=created))
            continue

        inv = str(invoice)
        if inv not in orders:
            if len(orders) >= MAX_ORDERS:
                capped = True
                break
            # customer
            cid_key = str(cust) if cust is not None else "guest"
            if cid_key not in customers:
                customers[cid_key] = next_cust_id
                next_cust_id += 1
            orders[inv] = Order(id=next_order_id, customer_id=customers[cid_key],
                                status="paid", total=0.0, created_at=created)
            next_order_id += 1
        o = orders[inv]
        items.append(OrderItem(order_id=o.id, product_id=0, sku=sku, qty=qty, price=round(price, 2)))
        o.total = round(o.total + qty * price, 2)

    wb.close()
    print(f"解析: products={len(products)} orders={len(orders)} items={len(items)} "
          f"returns={len(returns)} {'(已达上限截断)' if capped else ''}")

    # 写库
    db.add_all(list(products.values()))
    db.flush()
    pid = {p.sku: p.id for p in products.values()}
    # 客户
    cust_rows = [Customer(id=i, name=f"Customer {k}", email=f"c{i}@example.com",
                          tier=random.choice(["standard", "standard", "vip"]))
                 for k, i in customers.items()]
    db.add_all(cust_rows)
    db.add_all(list(orders.values()))
    db.flush()
    for it in items:
        it.product_id = pid.get(it.sku, 0)
    db.add_all(items)
    # 库存 (合成) + 广告 (合成) + 少量工单
    db.add_all([Inventory(sku=s, stock=random.randint(15, 500), safety_stock=30)
                for s in products])
    skus = list(products.keys())
    db.add_all([Campaign(name=f"campaign-{i}", sku=random.choice(skus),
                         channel=random.choice(["search", "social"]),
                         status=random.choice(["active", "active", "paused"]),
                         daily_budget=round(random.uniform(100, 1000), 2)) for i in range(50)])
    db.add_all([SupportTicket(sku=random.choice(skus), subject="咨询", body="(合成)",
                              status="open", created_at=NOW - timedelta(days=random.uniform(0, 20)))
                for _ in range(200)])
    db.add_all(returns)
    db.commit()
    print("真实交易数据写入完成 (库存/广告/工单为合成)")


def inject_anomaly(db) -> None:
    """重新注入演示用异常 SKU 与订单 10086, 保证 anomaly/refund/dashboard 演示可用。"""
    p = Product(sku=ANOMALY_SKU, name="Nexora 空气炸锅 Pro", category="kitchen",
                price=259.0, status="active")
    db.add(p)
    db.flush()
    db.add(Inventory(sku=ANOMALY_SKU, stock=12, safety_stock=30))
    db.add(Campaign(name=f"{ANOMALY_SKU} 夏促", sku=ANOMALY_SKU, channel="search",
                    status="active", daily_budget=800.0))
    cust = db.query(Customer.id).first()
    cid = cust[0] if cust else None

    next_oid = (db.query(Order.id).order_by(Order.id.desc()).first() or [0])[0] + 1
    def add_orders(n, lo, hi, oid_start):
        oid = oid_start
        rows_o, rows_i = [], []
        for _ in range(n):
            created = NOW - timedelta(days=random.uniform(lo, hi))
            rows_o.append(Order(id=oid, customer_id=cid, status="paid", total=259.0, created_at=created))
            rows_i.append(OrderItem(order_id=oid, product_id=p.id, sku=ANOMALY_SKU, qty=1, price=259.0))
            oid += 1
        return rows_o, rows_i, oid
    o1, i1, next_oid = add_orders(100, 7, 14, next_oid)
    o2, i2, next_oid = add_orders(62, 0, 7, next_oid)
    o3, i3, next_oid = add_orders(88, 14, 30, next_oid)
    all_o = o1 + o2 + o3
    db.add_all(all_o); db.add_all(i1 + i2 + i3); db.flush()
    for oid in random.sample([o.id for o in all_o], 46):
        db.add(Return(order_id=oid, sku=ANOMALY_SKU, reason="broken handle", amount=259.0,
                      created_at=NOW - timedelta(days=random.uniform(0, 10))))
    for _ in range(30):
        db.add(SupportTicket(sku=ANOMALY_SKU, subject="broken handle",
                             body="把手断裂 (broken handle)", status="open",
                             created_at=NOW - timedelta(days=random.uniform(0, 7))))
    # 订单 10086
    db.add(Order(id=10086, customer_id=cid, status="paid", total=300.0, created_at=NOW - timedelta(days=3)))
    db.add(OrderItem(order_id=10086, product_id=p.id, sku=ANOMALY_SKU, qty=1, price=300.0))
    db.commit()
    print("已注入演示异常 SKU + 订单 10086")


def main() -> None:
    db = SessionLocal()
    try:
        truncate(db)
        run_import(db)
        inject_anomaly(db)
    finally:
        db.close()


if __name__ == "__main__":
    main()
